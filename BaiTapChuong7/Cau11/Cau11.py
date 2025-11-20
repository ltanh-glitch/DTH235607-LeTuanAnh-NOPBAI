from __future__ import annotations
from pathlib import Path
from typing import List, Dict

from openpyxl import Workbook, load_workbook

EMPLOYEE_FILE = Path(__file__).with_name("employees.xlsx")

Employee = Dict[str, object]


def load_employees() -> List[Employee]:
    if not EMPLOYEE_FILE.exists():
        return []
    wb = load_workbook(EMPLOYEE_FILE, data_only=True)
    ws = wb.active
    employees: List[Employee] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        _, code, name, age = row
        if code is None:
            continue
        employees.append({
            "code": str(code).strip(),
            "name": str(name).strip() if name is not None else "",
            "age": int(age) if age is not None else 0,
        })
    return employees


def save_employees(employees: List[Employee]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nhân viên"
    ws.append(["STT", "Mã", "Tên", "Tuổi"])
    for idx, employee in enumerate(employees, 1):
        ws.append([idx, employee["code"], employee["name"], employee["age"]])
    wb.save(EMPLOYEE_FILE)
    print(f"Đã lưu {len(employees)} nhân viên vào {EMPLOYEE_FILE.name}")


def print_table(employees: List[Employee]) -> None:
    if not employees:
        print("Chưa có nhân viên nào.")
        return
    rows = [["STT", "Mã", "Tên", "Tuổi"]]
    for idx, employee in enumerate(employees, 1):
        rows.append([str(idx), employee["code"], employee["name"], str(employee["age"])])
    widths = [0] * len(rows[0])
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(cell))
    for row in rows:
        line = "  ".join(f"{cell:<{widths[i]}}" for i, cell in enumerate(row))
        print(line)


def prompt_nonempty(prompt: str) -> str:
    while True:
        value = input(prompt).strip()
        if value:
            return value
        print("Không được để trống.")


def prompt_age(prompt: str) -> int:
    while True:
        value = input(prompt).strip()
        if value.isdigit():
            return int(value)
        print("Tuổi phải là số nguyên dương.")


def find_employee(employees: List[Employee], code: str) -> Employee | None:
    for employee in employees:
        if employee["code"].lower() == code.lower():
            return employee
    return None


def sort_by_age(employees: List[Employee]) -> None:
    employees.sort(key=lambda emp: emp["age"])
    print("Đã sắp xếp nhân viên theo tuổi tăng dần.")


def main() -> None:
    employees = load_employees()
    if employees:
        print(f"Tải {len(employees)} nhân viên từ {EMPLOYEE_FILE.name}.")
    menu = {
        "1": "Thêm nhân viên",
        "2": "Hiển thị danh sách",
        "3": "Sắp xếp theo tuổi",
        "4": "Lưu vào Excel",
        "5": "Đọc Excel",
        "0": "Thoát",
    }
    while True:
        print("\n--- Quản lý nhân viên ---")
        for key, label in menu.items():
            print(f"{key}. {label}")
        choice = input("Chọn chức năng: ").strip()
        match choice:
            case "1":
                code = prompt_nonempty("Mã nhân viên: ")
                if find_employee(employees, code):
                    print("Mã đã tồn tại.")
                    continue
                name = prompt_nonempty("Tên nhân viên: ")
                age = prompt_age("Tuổi: ")
                employees.append({"code": code, "name": name, "age": age})
                print("Đã thêm nhân viên.")
            case "2":
                print_table(employees)
            case "3":
                if not employees:
                    print("Chưa có nhân viên để sắp xếp.")
                    continue
                sort_by_age(employees)
            case "4":
                save_employees(employees)
            case "5":
                employees = load_employees()
                print(f"Đã đọc {len(employees)} nhân viên từ file.")
            case "0":
                print("Kết thúc chương trình.")
                break
            case _:  # noqa: E999
                print("Chức năng không hợp lệ.")


if __name__ == "__main__":
    main()
