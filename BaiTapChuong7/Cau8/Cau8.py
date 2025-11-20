from pathlib import Path
from openpyxl import load_workbook

DATA_FILE = Path(__file__).with_name("demo.xlsx")
if not DATA_FILE.exists():
    alt = Path(__file__).parent.parent / "Cau7" / "demo.xlsx"
    DATA_FILE = alt
    if not DATA_FILE.exists():
        raise FileNotFoundError("demo.xlsx not found; run Cau7/Cau7.py first")
wb = load_workbook(DATA_FILE)
print(wb.sheetnames)

ws = wb[wb.sheetnames[0]]
rows = list(ws.values)
if not rows:
    raise SystemExit("Sheet is empty")

column_widths = [0] * len(rows[0])
for row in rows:
    for idx, value in enumerate(row):
        column_widths[idx] = max(column_widths[idx], len(str(value)) if value is not None else 0)

for row in rows:
    line = "".join(f"{'' if value is None else str(value):<{column_widths[idx] + 2}}" for idx, value in enumerate(row))
    print(line.rstrip())
